from __future__ import annotations

import csv
import json
import os
import tomllib
import unicodedata
from dataclasses import dataclass, field, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence

from ..ui.runtime_preflight.constants import RUNTIME_AGGREGATION_LOAD
from ..ui.runtime_preflight.models import RuntimeSelection
from ..ui.runtime_preflight.normalize import normalize_runtime_selection


TEXT_CONFIG_SCHEMA_VERSION = 1
TEXT_CONFIG_DIR_NAME = "pipeline29_text"
APP_STATE_DIR_NAME = "pipeline_newgen_rev1"
RUNTIME_STATE_FILENAME = "pipeline30_runtime_paths.json"

METADATA_FILENAME = "metadata.toml"
DEFAULTS_FILENAME = "defaults.toml"
DATA_QUALITY_FILENAME = "data_quality.toml"
MAPPINGS_FILENAME = "mappings.toml"
INSTRUMENTS_FILENAME = "instruments.toml"
REPORTING_FILENAME = "reporting_rounding.toml"
PLOTS_FILENAME = "plots.toml"
COMPARE_FILENAME = "compare.toml"
FUEL_PROPERTIES_FILENAME = "fuel_properties.toml"
KNOCK_THRESHOLDS_FILENAME = "knock_thresholds.toml"

DEFAULT_INSTRUMENT_COLUMNS = [
    "key",
    "component",
    "dist",
    "range_min",
    "range_max",
    "acc_abs",
    "acc_pct",
    "digits",
    "lsd",
    "resolution",
    "source",
    "notes",
    "setting_param",
    "setting_value",
]
DEFAULT_REPORTING_COLUMNS = ["key", "report_resolution", "report_digits", "rule", "notes"]
DEFAULT_PLOT_COLUMNS = [
    "enabled",
    "with_uncertainty",
    "without_uncertainty",
    "plot_type",
    "filename",
    "title",
    "x_col",
    "y_col",
    "yerr_col",
    "show_uncertainty",
    "x_label",
    "y_label",
    "x_min",
    "x_max",
    "x_step",
    "y_min",
    "y_max",
    "y_step",
    "y_tol_plus",
    "y_tol_minus",
    "filter_h2o_list",
    "label_variant",
    "notes",
]
DEFAULT_COMPARE_COLUMNS = [
    "enabled",
    "with_uncertainty",
    "without_uncertainty",
    "left_series",
    "right_series",
    "metric_id",
    "show_uncertainty",
    "notes",
]
DEFAULT_FUEL_PROPERTY_COLUMNS = [
    "Fuel_Label",
    "DIES_pct",
    "BIOD_pct",
    "EtOH_pct",
    "H2O_pct",
    "LHV_kJ_kg",
    "Fuel_Density_kg_m3",
    "Fuel_Cost_R_L",
    "reference",
    "notes",
]
DEFAULT_KNOCK_THRESHOLD_COLUMNS = ["threshold_bar", "enabled", "notes"]
REQUIRED_MAPPING_KEYS = {"power_kw", "fuel_kgh", "lhv_kj_kg"}


@dataclass
class ConfigBundle:
    mappings: Dict[str, Dict[str, str]]
    instruments: List[Dict[str, Any]]
    reporting: List[Dict[str, Any]]
    plots: List[Dict[str, Any]]
    compare: List[Dict[str, Any]]
    fuel_properties: List[Dict[str, Any]]
    data_quality: Dict[str, float]
    knock_thresholds: List[Dict[str, Any]] = field(default_factory=list)
    defaults: Dict[str, str] = field(default_factory=dict)
    source_kind: str = "text"
    source_path: Optional[Path] = None
    text_dir: Optional[Path] = None
    metadata: Dict[str, Any] = field(default_factory=dict)


@dataclass
class RuntimeState:
    raw_input_dir: Optional[Path] = None
    out_dir: Optional[Path] = None
    selection: RuntimeSelection = field(
        default_factory=lambda: RuntimeSelection(aggregation_mode=RUNTIME_AGGREGATION_LOAD)
    )
    helper_configured: bool = False
    dirs_configured_in_gui: bool = False
    config_dir: Optional[Path] = None
    extra: Dict[str, Any] = field(default_factory=dict)


def default_text_config_dir(project_root: Path) -> Path:
    return Path(project_root).expanduser().resolve() / "config" / TEXT_CONFIG_DIR_NAME


def default_legacy_excel_path(project_root: Path) -> Path:
    return Path(project_root).expanduser().resolve() / "config" / "config_incertezas_rev3.xlsx"


def default_app_state_dir() -> Path:
    return Path(os.environ.get("LOCALAPPDATA", str(Path.home()))).expanduser().resolve() / APP_STATE_DIR_NAME


def default_runtime_state_path() -> Path:
    return default_app_state_dir() / RUNTIME_STATE_FILENAME


def bundle_required_paths(config_dir: Path) -> Dict[str, Path]:
    config_dir = Path(config_dir).expanduser().resolve()
    return {
        "metadata": config_dir / METADATA_FILENAME,
        "defaults": config_dir / DEFAULTS_FILENAME,
        "data_quality": config_dir / DATA_QUALITY_FILENAME,
        "mappings": config_dir / MAPPINGS_FILENAME,
        "instruments": config_dir / INSTRUMENTS_FILENAME,
        "reporting": config_dir / REPORTING_FILENAME,
        "plots": config_dir / PLOTS_FILENAME,
        "compare": config_dir / COMPARE_FILENAME,
        "fuel_properties": config_dir / FUEL_PROPERTIES_FILENAME,
        "knock_thresholds": config_dir / KNOCK_THRESHOLDS_FILENAME,
    }


def text_config_exists(config_dir: Path) -> bool:
    required_names = {"metadata", "defaults", "data_quality", "mappings", "instruments", "reporting", "plots"}
    paths = bundle_required_paths(config_dir)
    return all(path.exists() for name, path in paths.items() if name in required_names)


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    return str(value).replace("\ufeff", "").strip() == ""


def _to_text(value: Any) -> str:
    if _is_blank(value):
        return ""
    return str(value).replace("\ufeff", "").strip()


def _normalize_key(value: Any) -> str:
    text = _to_text(value).lower()
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _to_text(value).lower() in {"1", "true", "yes", "on", "y", "checked"}


def _to_optional_path(value: Any) -> Optional[Path]:
    text = _to_text(value)
    if not text:
        return None
    return Path(text).expanduser().resolve()


def _to_builtin_scalar(value: Any) -> Any:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return int(value)
    if isinstance(value, float):
        return float(value)
    text = _to_text(value)
    return text if text else None


def _clean_record(record: Mapping[str, Any], preferred_order: Sequence[str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    seen: set[str] = set()
    for key in preferred_order:
        if key not in record:
            continue
        clean = _to_builtin_scalar(record.get(key))
        if clean is None:
            continue
        out[str(key)] = clean
        seen.add(str(key))
    for key, value in record.items():
        key_text = _to_text(key)
        if not key_text or key_text in seen:
            continue
        clean = _to_builtin_scalar(value)
        if clean is None:
            continue
        out[key_text] = clean
    return out


def _normalize_rows(rows: Sequence[Mapping[str, Any]] | None, preferred_order: Sequence[str]) -> List[Dict[str, Any]]:
    if not rows:
        return []
    normalized: List[Dict[str, Any]] = []
    for row in rows:
        cleaned = _clean_record(dict(row), preferred_order)
        if cleaned:
            normalized.append(cleaned)
    return normalized


def _read_toml_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    return tomllib.loads(path.read_text(encoding="utf-8"))


def _toml_scalar(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        return repr(float(value))
    return json.dumps(str(value), ensure_ascii=False)


def _write_toml_table_file(path: Path, table_name: str, values: Mapping[str, Any]) -> None:
    lines = [f"schema_version = {TEXT_CONFIG_SCHEMA_VERSION}", "", f"[{table_name}]"]
    for key, value in values.items():
        clean = _to_builtin_scalar(value)
        if clean is None:
            continue
        lines.append(f"{json.dumps(str(key), ensure_ascii=False)} = {_toml_scalar(clean)}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_toml_keyed_tables(
    path: Path,
    table_name: str,
    values: Mapping[str, Mapping[str, Any]],
    field_order: Sequence[str],
) -> None:
    lines = [f"schema_version = {TEXT_CONFIG_SCHEMA_VERSION}", ""]
    for key, record in values.items():
        clean = _clean_record(record, field_order)
        if not clean:
            continue
        lines.append(f"[{table_name}.{json.dumps(str(key), ensure_ascii=False)}]")
        for field, value in clean.items():
            lines.append(f"{json.dumps(str(field), ensure_ascii=False)} = {_toml_scalar(value)}")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_toml_array_of_tables(path: Path, array_name: str, rows: Sequence[Mapping[str, Any]], field_order: Sequence[str]) -> None:
    lines = [f"schema_version = {TEXT_CONFIG_SCHEMA_VERSION}", ""]
    for row in rows:
        clean = _clean_record(row, field_order)
        if not clean:
            continue
        lines.append(f"[[{array_name}]]")
        for key, value in clean.items():
            lines.append(f"{json.dumps(str(key), ensure_ascii=False)} = {_toml_scalar(value)}")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def save_text_config_bundle(
    bundle: ConfigBundle,
    config_dir: Path,
    *,
    bootstrapped_from: Optional[Path] = None,
) -> ConfigBundle:
    config_dir = Path(config_dir).expanduser().resolve()
    config_dir.mkdir(parents=True, exist_ok=True)
    metadata = {
        "format": "pipeline29_text",
        "schema_version": TEXT_CONFIG_SCHEMA_VERSION,
        "updated_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
    }
    if bootstrapped_from is not None:
        metadata["bootstrapped_from"] = str(Path(bootstrapped_from).expanduser().resolve())
    if bundle.source_kind:
        metadata["source_kind"] = bundle.source_kind
    if bundle.source_path is not None:
        metadata["source_path"] = str(Path(bundle.source_path).expanduser().resolve())

    paths = bundle_required_paths(config_dir)
    _write_toml_table_file(paths["metadata"], "metadata", metadata)
    _write_toml_table_file(paths["defaults"], "defaults", dict(bundle.defaults))
    _write_toml_table_file(paths["data_quality"], "data_quality", dict(bundle.data_quality))
    _write_toml_keyed_tables(paths["mappings"], "mappings", dict(bundle.mappings), ["mean", "sd", "unit", "notes"])
    _write_toml_array_of_tables(paths["instruments"], "instruments", bundle.instruments, DEFAULT_INSTRUMENT_COLUMNS)
    _write_toml_array_of_tables(paths["reporting"], "reporting", bundle.reporting, DEFAULT_REPORTING_COLUMNS)
    _write_toml_array_of_tables(paths["plots"], "plots", bundle.plots, DEFAULT_PLOT_COLUMNS)
    _write_toml_array_of_tables(paths["compare"], "compare", bundle.compare, DEFAULT_COMPARE_COLUMNS)
    _write_toml_array_of_tables(
        paths["fuel_properties"],
        "fuel_properties",
        bundle.fuel_properties,
        DEFAULT_FUEL_PROPERTY_COLUMNS,
    )
    _write_toml_array_of_tables(
        paths["knock_thresholds"],
        "knock_thresholds",
        bundle.knock_thresholds,
        DEFAULT_KNOCK_THRESHOLD_COLUMNS,
    )
    return replace(
        bundle,
        source_kind="text",
        source_path=config_dir,
        text_dir=config_dir,
        metadata=metadata,
    )


def load_text_config_bundle(config_dir: Path) -> ConfigBundle:
    config_dir = Path(config_dir).expanduser().resolve()
    paths = bundle_required_paths(config_dir)
    if not text_config_exists(config_dir):
        missing = [str(path.name) for name, path in paths.items() if name != "compare" and name != "fuel_properties" and not path.exists()]
        raise FileNotFoundError(f"Config textual incompleta em {config_dir}: faltam {missing}")

    metadata_doc = _read_toml_file(paths["metadata"])
    defaults_doc = _read_toml_file(paths["defaults"])
    data_quality_doc = _read_toml_file(paths["data_quality"])
    mappings_doc = _read_toml_file(paths["mappings"])
    instruments_doc = _read_toml_file(paths["instruments"])
    reporting_doc = _read_toml_file(paths["reporting"])
    plots_doc = _read_toml_file(paths["plots"])
    compare_doc = _read_toml_file(paths["compare"])
    fuel_properties_doc = _read_toml_file(paths["fuel_properties"])
    knock_thresholds_doc = _read_toml_file(paths["knock_thresholds"])

    metadata = dict(metadata_doc.get("metadata", {}) or {})
    if "schema_version" not in metadata:
        metadata["schema_version"] = metadata_doc.get("schema_version", TEXT_CONFIG_SCHEMA_VERSION)

    defaults = {str(key): _to_text(value) for key, value in (defaults_doc.get("defaults", {}) or {}).items()}
    data_quality: Dict[str, float] = {}
    for key, value in (data_quality_doc.get("data_quality", {}) or {}).items():
        try:
            data_quality[str(key)] = float(value)
        except Exception:
            continue

    mappings: Dict[str, Dict[str, str]] = {}
    for key, spec in (mappings_doc.get("mappings", {}) or {}).items():
        key_text = _to_text(key)
        if not key_text:
            continue
        row = dict(spec or {})
        mappings[key_text] = {
            "mean": _to_text(row.get("mean", "")),
            "sd": _to_text(row.get("sd", "")),
            "unit": _to_text(row.get("unit", "")),
            "notes": _to_text(row.get("notes", "")),
        }

    source_path = _to_optional_path(metadata.get("source_path")) or config_dir
    return ConfigBundle(
        mappings=mappings,
        instruments=_normalize_rows(instruments_doc.get("instruments", []) or [], DEFAULT_INSTRUMENT_COLUMNS),
        reporting=_normalize_rows(reporting_doc.get("reporting", []) or [], DEFAULT_REPORTING_COLUMNS),
        plots=_normalize_rows(plots_doc.get("plots", []) or [], DEFAULT_PLOT_COLUMNS),
        compare=_normalize_rows(compare_doc.get("compare", []) or [], DEFAULT_COMPARE_COLUMNS),
        fuel_properties=_normalize_rows(
            fuel_properties_doc.get("fuel_properties", []) or [],
            DEFAULT_FUEL_PROPERTY_COLUMNS,
        ),
        knock_thresholds=_normalize_rows(
            knock_thresholds_doc.get("knock_thresholds", []) or [],
            DEFAULT_KNOCK_THRESHOLD_COLUMNS,
        ),
        data_quality=data_quality,
        defaults=defaults,
        source_kind="text",
        source_path=source_path,
        text_dir=config_dir,
        metadata=metadata,
    )


def validate_bundle(bundle: ConfigBundle) -> List[str]:
    mapping_keys = {_normalize_key(key) for key in bundle.mappings}
    missing_keys = sorted(REQUIRED_MAPPING_KEYS - mapping_keys)
    if not missing_keys:
        return []
    return [f"Mappings sem chaves obrigatorias: {missing_keys}"]


def summarize_config_bundle(bundle: ConfigBundle) -> Dict[str, Any]:
    return {
        "source_kind": bundle.source_kind,
        "source_path": str(bundle.source_path) if bundle.source_path is not None else "",
        "text_dir": str(bundle.text_dir) if bundle.text_dir is not None else "",
        "mapping_count": len(bundle.mappings),
        "instrument_rows": len(bundle.instruments),
        "reporting_rows": len(bundle.reporting),
        "plot_rows": len(bundle.plots),
        "compare_rows": len(bundle.compare),
        "fuel_property_rows": len(bundle.fuel_properties),
        "defaults_keys": len(bundle.defaults),
        "data_quality_keys": len(bundle.data_quality),
        "validation_errors": validate_bundle(bundle),
    }


def _require_openpyxl() -> Any:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(
            "Legacy Excel config support requires the optional dependency 'openpyxl'."
        ) from exc
    return load_workbook


def _worksheet_rows(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    load_workbook = _require_openpyxl()
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        selected = None
        for candidate in workbook.sheetnames:
            if candidate == sheet_name or _normalize_key(candidate) == _normalize_key(sheet_name):
                selected = workbook[candidate]
                break
        if selected is None:
            return []
        rows = list(selected.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []
    header = [_to_text(value) for value in rows[0]]
    out: List[Dict[str, Any]] = []
    for raw_row in rows[1:]:
        if raw_row is None:
            continue
        row: Dict[str, Any] = {}
        for index, field_name in enumerate(header):
            if not field_name:
                continue
            row[field_name] = raw_row[index] if index < len(raw_row) else None
        if any(not _is_blank(value) for value in row.values()):
            out.append(row)
    return out


def _sniff_csv_dialect(sample: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t")
    except Exception:
        return csv.get_dialect("excel")


def _format_pct_for_label(value: Any) -> str:
    try:
        numeric = float(value)
    except Exception:
        return _to_text(value)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:g}"


def _infer_fuel_label(dies_pct: Any, biod_pct: Any, etoh_pct: Any, h2o_pct: Any) -> str:
    try:
        dies = float(dies_pct)
        biod = float(biod_pct)
        etoh = float(etoh_pct)
        h2o = float(h2o_pct)
    except Exception:
        return ""

    def near_zero(value: float) -> bool:
        return abs(value) < 1e-9

    if near_zero(etoh) and near_zero(h2o):
        if near_zero(dies):
            return f"B{_format_pct_for_label(biod)}"
        if near_zero(biod):
            return f"D{_format_pct_for_label(dies)}"
        return f"D{_format_pct_for_label(dies)}B{_format_pct_for_label(biod)}"
    if near_zero(dies) and near_zero(biod):
        if near_zero(h2o):
            return f"E{_format_pct_for_label(etoh)}"
        if near_zero(etoh):
            return f"H{_format_pct_for_label(h2o)}"
        return f"E{_format_pct_for_label(etoh)}H{_format_pct_for_label(h2o)}"
    if near_zero(dies) and near_zero(h2o):
        return f"B{_format_pct_for_label(biod)}E{_format_pct_for_label(etoh)}"
    if near_zero(biod) and near_zero(h2o):
        return f"D{_format_pct_for_label(dies)}E{_format_pct_for_label(etoh)}"
    return ""


def _lookup_default_value(defaults: Mapping[str, str], key_name: str) -> str:
    wanted = _normalize_key(key_name)
    for key, value in defaults.items():
        if _normalize_key(key) == wanted:
            return _to_text(value)
    return ""


def _legacy_fuel_properties_from_csv(csv_path: Path, defaults: Mapping[str, str]) -> List[Dict[str, Any]]:
    if not csv_path.exists():
        return []
    sample = csv_path.read_text(encoding="utf-8-sig")
    dialect = _sniff_csv_dialect(sample[:4096])
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, dialect=dialect)
        records = list(reader)

    out: List[Dict[str, Any]] = []
    for raw_row in records:
        row = {str(key): value for key, value in raw_row.items()}
        normalized_row: Dict[str, Any] = {}
        for column, value in row.items():
            column_norm = _normalize_key(column)
            if column_norm in {"fuel_label", "label"}:
                normalized_row["Fuel_Label"] = value
            elif column_norm in {"dies_pct", "dies", "diesel_pct", "diesel"}:
                normalized_row["DIES_pct"] = value
            elif column_norm in {"biod_pct", "biod", "biodiesel_pct", "biodiesel"}:
                normalized_row["BIOD_pct"] = value
            elif column_norm in {"etoh_pct", "etoh", "e_pct", "e"}:
                normalized_row["EtOH_pct"] = value
            elif column_norm in {"h2o_pct", "h20_pct", "h2o", "h20", "h_pct", "h"}:
                normalized_row["H2O_pct"] = value
            elif column_norm in {"lhv_kj_kg", "lhv", "pci_kj_kg", "pci"}:
                normalized_row["LHV_kJ_kg"] = value
            elif column_norm in {"fuel_density_kg_m3", "density_kg_m3", "density"}:
                normalized_row["Fuel_Density_kg_m3"] = value
            elif column_norm in {"fuel_cost_r_l", "cost_r_l", "cost"}:
                normalized_row["Fuel_Cost_R_L"] = value
            elif column_norm in {"reference", "source"}:
                normalized_row["reference"] = value
            elif column_norm in {"notes", "note"}:
                normalized_row["notes"] = value

        label = _to_text(normalized_row.get("Fuel_Label", ""))
        if not label:
            label = _infer_fuel_label(
                normalized_row.get("DIES_pct", ""),
                normalized_row.get("BIOD_pct", ""),
                normalized_row.get("EtOH_pct", ""),
                normalized_row.get("H2O_pct", ""),
            )
        if not label:
            continue

        density = _to_text(normalized_row.get("Fuel_Density_kg_m3", ""))
        cost = _to_text(normalized_row.get("Fuel_Cost_R_L", ""))
        if not density:
            density = _lookup_default_value(defaults, f"FUEL_DENSITY_KG_M3_{label}")
        if not cost:
            cost = _lookup_default_value(defaults, f"FUEL_COST_R_L_{label}")

        out.append(
            {
                "Fuel_Label": label,
                "DIES_pct": _to_text(normalized_row.get("DIES_pct", "")),
                "BIOD_pct": _to_text(normalized_row.get("BIOD_pct", "")),
                "EtOH_pct": _to_text(normalized_row.get("EtOH_pct", "")),
                "H2O_pct": _to_text(normalized_row.get("H2O_pct", "")),
                "LHV_kJ_kg": _to_text(normalized_row.get("LHV_kJ_kg", "")),
                "Fuel_Density_kg_m3": density,
                "Fuel_Cost_R_L": cost,
                "reference": _to_text(normalized_row.get("reference", "")) or f"Imported from {csv_path.name}",
                "notes": _to_text(normalized_row.get("notes", "")),
            }
        )
    return _normalize_rows(out, DEFAULT_FUEL_PROPERTY_COLUMNS)


def load_excel_config_bundle(excel_path: Path) -> ConfigBundle:
    excel_path = Path(excel_path).expanduser().resolve()
    mapping_rows = _worksheet_rows(excel_path, "Mappings")
    mappings: Dict[str, Dict[str, str]] = {}
    for row in mapping_rows:
        key = _to_text(row.get("key", ""))
        mean = _to_text(row.get("col_mean", ""))
        if not key or "logical variable identifier" in key.lower():
            continue
        if "exact dataframe column name" in mean.lower():
            continue
        mappings[key] = {
            "mean": mean,
            "sd": _to_text(row.get("col_sd", "")),
            "unit": _to_text(row.get("unit", "")),
            "notes": _to_text(row.get("notes", "")),
        }

    defaults: Dict[str, str] = {}
    for row in _worksheet_rows(excel_path, "Defaults"):
        param = _to_text(row.get("param", ""))
        if not param or "global parameter name" in param.lower():
            continue
        defaults[param] = _to_text(row.get("value", ""))

    data_quality: Dict[str, float] = {}
    for row in _worksheet_rows(excel_path, "data quality assessment"):
        param = _to_text(row.get("param", ""))
        if not param:
            continue
        try:
            data_quality[param] = float(row.get("value", ""))
        except Exception:
            continue

    plots = _normalize_rows(_worksheet_rows(excel_path, "Plots"), DEFAULT_PLOT_COLUMNS)
    for row in plots:
        if _to_text(row.get("show_uncertainty", "")):
            continue
        yerr = _to_text(row.get("yerr_col", "")).lower()
        row["show_uncertainty"] = "off" if yerr in {"off", "none", "disabled", "disable", "0", "na", "n/a"} else "auto"

    reporting_rows = _worksheet_rows(excel_path, "Reporting_Rounding")
    if not reporting_rows:
        reporting_rows = _worksheet_rows(excel_path, "UPD_Rounding")

    return ConfigBundle(
        mappings=mappings,
        instruments=_normalize_rows(_worksheet_rows(excel_path, "Instruments"), DEFAULT_INSTRUMENT_COLUMNS),
        reporting=_normalize_rows(reporting_rows, DEFAULT_REPORTING_COLUMNS),
        plots=plots,
        compare=[],
        fuel_properties=_legacy_fuel_properties_from_csv(excel_path.parent / "lhv.csv", defaults),
        data_quality=data_quality,
        defaults=defaults,
        source_kind="excel",
        source_path=excel_path,
        text_dir=None,
        metadata={"bootstrapped_from": str(excel_path)},
    )


def bootstrap_text_config_from_excel(excel_path: Path, config_dir: Path) -> ConfigBundle:
    bundle = load_excel_config_bundle(excel_path)
    return save_text_config_bundle(bundle, config_dir, bootstrapped_from=Path(excel_path).expanduser().resolve())


def load_pipeline29_config_bundle(
    *,
    project_root: Path,
    config_source: str = "auto",
    text_config_dir: Optional[Path] = None,
    rebuild_text_config: bool = False,
    excel_path: Optional[Path] = None,
) -> ConfigBundle:
    source_mode = _to_text(config_source).lower() or "auto"
    if source_mode not in {"auto", "text", "excel"}:
        raise ValueError(f"config_source invalido: {config_source}")

    text_dir = (
        Path(text_config_dir).expanduser().resolve()
        if text_config_dir is not None
        else default_text_config_dir(project_root)
    )
    chosen_excel_path = (
        Path(excel_path).expanduser().resolve()
        if excel_path is not None
        else default_legacy_excel_path(project_root)
    )

    if source_mode in {"auto", "text"}:
        if rebuild_text_config or not text_config_exists(text_dir):
            if chosen_excel_path.exists():
                bootstrap_text_config_from_excel(chosen_excel_path, text_dir)
            elif source_mode == "text":
                raise FileNotFoundError(f"Nao encontrei config textual completa em {text_dir}")
        if text_config_exists(text_dir):
            bundle = load_text_config_bundle(text_dir)
            return replace(bundle, source_kind="text", source_path=text_dir, text_dir=text_dir)
        if source_mode == "text":
            raise FileNotFoundError(f"Nao encontrei config textual completa em {text_dir}")

    if not chosen_excel_path.exists():
        raise FileNotFoundError(f"Nao encontrei {chosen_excel_path.name} em {chosen_excel_path.parent}")
    bundle = load_excel_config_bundle(chosen_excel_path)
    return replace(bundle, text_dir=text_dir if text_config_exists(text_dir) else None)


def runtime_state_from_mapping(payload: Mapping[str, Any]) -> RuntimeState:
    selection = normalize_runtime_selection(
        RuntimeSelection(
            aggregation_mode=payload.get("aggregation_mode", RUNTIME_AGGREGATION_LOAD),
            sweep_key=payload.get("sweep_key", ""),
            sweep_x_col=payload.get("sweep_x_col", ""),
            sweep_bin_tol=payload.get("sweep_bin_tol", 0.0),
        )
    )
    known_keys = {
        "raw_input_dir",
        "out_dir",
        "aggregation_mode",
        "sweep_key",
        "sweep_x_col",
        "sweep_bin_tol",
        "helper_configured",
        "dirs_configured_in_gui",
        "config_dir",
    }
    return RuntimeState(
        raw_input_dir=_to_optional_path(payload.get("raw_input_dir")),
        out_dir=_to_optional_path(payload.get("out_dir")),
        selection=selection,
        helper_configured=_to_bool(payload.get("helper_configured")),
        dirs_configured_in_gui=_to_bool(payload.get("dirs_configured_in_gui")),
        config_dir=_to_optional_path(payload.get("config_dir")),
        extra={str(key): value for key, value in payload.items() if str(key) not in known_keys},
    )


def runtime_state_as_mapping(state: RuntimeState) -> Dict[str, Any]:
    normalized = replace(state, selection=normalize_runtime_selection(state.selection))
    payload = dict(normalized.extra)
    if normalized.raw_input_dir is not None:
        payload["raw_input_dir"] = str(normalized.raw_input_dir)
    if normalized.out_dir is not None:
        payload["out_dir"] = str(normalized.out_dir)
    if normalized.config_dir is not None:
        payload["config_dir"] = str(normalized.config_dir)
    payload["aggregation_mode"] = normalized.selection.aggregation_mode
    payload["sweep_key"] = normalized.selection.sweep_key
    payload["sweep_x_col"] = normalized.selection.sweep_x_col
    payload["sweep_bin_tol"] = normalized.selection.sweep_bin_tol
    payload["helper_configured"] = bool(normalized.helper_configured)
    payload["dirs_configured_in_gui"] = bool(normalized.dirs_configured_in_gui)
    return payload


def load_runtime_state(path: Path) -> RuntimeState:
    path = Path(path).expanduser().resolve()
    if not path.exists():
        return RuntimeState()
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return RuntimeState()
    if not isinstance(payload, dict):
        return RuntimeState()
    return runtime_state_from_mapping(payload)


def save_runtime_state(path: Path, state: RuntimeState) -> RuntimeState:
    path = Path(path).expanduser().resolve()
    normalized = replace(state, selection=normalize_runtime_selection(state.selection))
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(runtime_state_as_mapping(normalized), indent=2, ensure_ascii=True), encoding="utf-8")
    return normalized


def summarize_runtime_state(state: RuntimeState, *, state_path: Optional[Path] = None) -> Dict[str, Any]:
    summary = {
        "raw_input_dir": str(state.raw_input_dir) if state.raw_input_dir is not None else "",
        "out_dir": str(state.out_dir) if state.out_dir is not None else "",
        "aggregation_mode": state.selection.aggregation_mode,
        "sweep_key": state.selection.sweep_key,
        "sweep_x_col": state.selection.sweep_x_col,
        "sweep_bin_tol": state.selection.sweep_bin_tol,
        "helper_configured": bool(state.helper_configured),
        "dirs_configured_in_gui": bool(state.dirs_configured_in_gui),
        "config_dir": str(state.config_dir) if state.config_dir is not None else "",
        "extra_keys": sorted(state.extra.keys()),
    }
    if state_path is not None:
        summary["state_path"] = str(Path(state_path).expanduser().resolve())
    return summary


def _resolve_runtime_dir(primary: Optional[Path], fallback_value: Any, fallback_default: Path) -> Path:
    if primary is not None:
        return primary
    fallback_path = _to_optional_path(fallback_value)
    if fallback_path is not None:
        return fallback_path
    return Path(fallback_default).expanduser().resolve()


def apply_runtime_path_overrides(
    defaults: Mapping[str, Any],
    runtime_state: Optional[RuntimeState] = None,
    *,
    default_process_dir: Optional[Path] = None,
    default_out_dir: Optional[Path] = None,
) -> tuple[Dict[str, str], Path, Path]:
    normalized_defaults = {
        _normalize_key(key): _to_text(value)
        for key, value in defaults.items()
        if _normalize_key(key)
    }
    state = runtime_state or RuntimeState()
    input_dir = _resolve_runtime_dir(
        state.raw_input_dir,
        normalized_defaults.get("raw_input_dir", ""),
        default_process_dir or Path.cwd(),
    )
    out_dir = _resolve_runtime_dir(
        state.out_dir,
        normalized_defaults.get("out_dir", ""),
        default_out_dir or input_dir,
    )
    normalized_defaults["raw_input_dir"] = str(input_dir)
    normalized_defaults["out_dir"] = str(out_dir)
    return normalized_defaults, input_dir, out_dir


def _write_runtime_dirs_to_defaults_excel(xlsx_path: Path, input_dir: Path, out_dir: Path) -> None:
    load_workbook = _require_openpyxl()
    workbook = load_workbook(xlsx_path)
    try:
        sheet = None
        for candidate in workbook.sheetnames:
            if candidate == "Defaults" or _normalize_key(candidate) == "defaults":
                sheet = workbook[candidate]
                break
        if sheet is None:
            raise RuntimeError(f"Nao encontrei a aba Defaults em {xlsx_path}")
        header = [_normalize_key(sheet.cell(row=1, column=index).value) for index in range(1, sheet.max_column + 1)]
        header_map = {value: index + 1 for index, value in enumerate(header) if value}
        param_col = header_map.get("param", 1)
        value_col = header_map.get("value", 2)
        replacements = {"raw_input_dir": str(Path(input_dir).expanduser().resolve()), "out_dir": str(Path(out_dir).expanduser().resolve())}
        updated: set[str] = set()
        for row_idx in range(1, sheet.max_row + 1):
            param_value = _normalize_key(sheet.cell(row=row_idx, column=param_col).value)
            if param_value in replacements:
                sheet.cell(row=row_idx, column=value_col).value = replacements[param_value]
                updated.add(param_value)
        next_row = sheet.max_row + 1
        for param_key, value in replacements.items():
            if param_key in updated:
                continue
            sheet.cell(row=next_row, column=param_col).value = "RAW_INPUT_DIR" if param_key == "raw_input_dir" else "OUT_DIR"
            sheet.cell(row=next_row, column=value_col).value = value
            next_row += 1
        workbook.save(xlsx_path)
    finally:
        workbook.close()


def sync_runtime_dirs_to_config_source(bundle: Optional[ConfigBundle], input_dir: Path, out_dir: Path) -> None:
    if bundle is None or bundle.source_path is None:
        return
    if bundle.source_kind == "excel":
        _write_runtime_dirs_to_defaults_excel(bundle.source_path, input_dir, out_dir)
