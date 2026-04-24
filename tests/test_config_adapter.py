from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

from _path import ROOT  # noqa: F401
from pipeline_newgen_rev1.config import (
    ConfigBundle,
    RuntimeState,
    apply_runtime_path_overrides,
    bootstrap_text_config_from_excel,
    load_pipeline29_config_bundle,
    load_runtime_state,
    load_text_config_bundle,
    save_runtime_state,
    save_text_config_bundle,
    summarize_config_bundle,
    summarize_runtime_state,
    text_config_exists,
)
from pipeline_newgen_rev1.ui.runtime_preflight.models import RuntimeSelection


def _sample_bundle() -> ConfigBundle:
    return ConfigBundle(
        mappings={
            "power_kw": {"mean": "Power_kW", "sd": "Power_SD", "unit": "kW", "notes": ""},
            "fuel_kgh": {"mean": "Fuel_kg_h", "sd": "", "unit": "kg/h", "notes": ""},
            "lhv_kj_kg": {"mean": "LHV_kJ_kg", "sd": "", "unit": "kJ/kg", "notes": ""},
        },
        instruments=[{"key": "power_kw", "component": "dyno", "source": "bench"}],
        reporting=[{"key": "power_kw", "report_resolution": "0.1", "report_digits": "1", "rule": "round"}],
        plots=[{"enabled": "1", "plot_type": "yx", "filename": "power_vs_load.png", "x_col": "Load_kW", "y_col": "Power_kW"}],
        compare=[],
        fuel_properties=[],
        data_quality={"MAX_DELTA_BETWEEN_SAMPLES_ms": 500.0},
        defaults={"RAW_INPUT_DIR": r"C:\input", "OUT_DIR": r"C:\out"},
    )


class ConfigAdapterTests(unittest.TestCase):
    def test_text_bundle_round_trip(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            config_dir = Path(tmpdir) / "pipeline29_text"
            saved = save_text_config_bundle(_sample_bundle(), config_dir)
            loaded = load_text_config_bundle(config_dir)
            summary = summarize_config_bundle(loaded)

            self.assertTrue(text_config_exists(config_dir))
            self.assertEqual(saved.source_kind, "text")
            self.assertEqual(loaded.defaults["RAW_INPUT_DIR"], r"C:\input")
            self.assertEqual(loaded.mappings["power_kw"]["mean"], "Power_kW")
            self.assertEqual(summary["mapping_count"], 3)
            self.assertEqual(summary["validation_errors"], [])

    def test_auto_loader_prefers_existing_text_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            root = Path(tmpdir)
            text_dir = root / "config" / "pipeline29_text"
            save_text_config_bundle(_sample_bundle(), text_dir)

            loaded = load_pipeline29_config_bundle(project_root=root, config_source="auto")

            self.assertEqual(loaded.source_kind, "text")
            self.assertEqual(loaded.source_path, text_dir.resolve())
            self.assertEqual(loaded.text_dir, text_dir.resolve())

    def test_runtime_state_round_trip_and_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            state_path = Path(tmpdir) / "runtime.json"
            state = RuntimeState(
                raw_input_dir=Path(tmpdir) / "raw",
                out_dir=Path(tmpdir) / "out",
                selection=RuntimeSelection(
                    aggregation_mode="sweep",
                    sweep_key="lambda",
                    sweep_x_col="Lambda_Medida",
                    sweep_bin_tol=0.02,
                ),
                helper_configured=True,
                dirs_configured_in_gui=True,
                config_dir=Path(tmpdir) / "cfg",
                extra={"custom_flag": "x"},
            )

            saved = save_runtime_state(state_path, state)
            loaded = load_runtime_state(state_path)
            summary = summarize_runtime_state(loaded, state_path=state_path)

            self.assertEqual(saved.selection.aggregation_mode, "sweep")
            self.assertEqual(loaded.selection.sweep_key, "lambda")
            self.assertTrue(loaded.helper_configured)
            self.assertTrue(loaded.dirs_configured_in_gui)
            self.assertEqual(summary["extra_keys"], ["custom_flag"])

    def test_apply_runtime_path_overrides_uses_saved_state_first(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            state = RuntimeState(
                raw_input_dir=Path(tmpdir) / "raw_saved",
                out_dir=Path(tmpdir) / "out_saved",
                selection=RuntimeSelection(aggregation_mode="load"),
            )
            defaults = {"RAW_INPUT_DIR": r"C:\default_input", "OUT_DIR": r"C:\default_out", "OTHER": "keep"}

            merged, input_dir, out_dir = apply_runtime_path_overrides(defaults, state)

            self.assertEqual(input_dir, (Path(tmpdir) / "raw_saved").resolve())
            self.assertEqual(out_dir, (Path(tmpdir) / "out_saved").resolve())
            self.assertEqual(merged["raw_input_dir"], str((Path(tmpdir) / "raw_saved").resolve()))
            self.assertEqual(merged["out_dir"], str((Path(tmpdir) / "out_saved").resolve()))
            self.assertEqual(merged["other"], "keep")

    def test_missing_runtime_state_file_returns_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            state = load_runtime_state(Path(tmpdir) / "missing.json")
            self.assertEqual(state.selection.aggregation_mode, "load")
            self.assertIsNone(state.raw_input_dir)
            self.assertIsNone(state.out_dir)

    def test_excel_bootstrap_is_optional(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            excel_path = Path(tmpdir) / "config_incertezas_rev3.xlsx"
            text_dir = Path(tmpdir) / "text_bundle"

            if importlib.util.find_spec("openpyxl") is None:
                with self.assertRaises(RuntimeError):
                    bootstrap_text_config_from_excel(excel_path, text_dir)
                return

            from openpyxl import Workbook

            workbook = Workbook()
            ws = workbook.active
            ws.title = "Mappings"
            ws.append(["key", "col_mean", "col_sd", "unit", "notes"])
            ws.append(["power_kw", "Power_kW", "", "kW", ""])
            ws.append(["fuel_kgh", "Fuel_kg_h", "", "kg/h", ""])
            ws.append(["lhv_kj_kg", "LHV_kJ_kg", "", "kJ/kg", ""])

            defaults_ws = workbook.create_sheet("Defaults")
            defaults_ws.append(["param", "value", "notes"])
            defaults_ws.append(["RAW_INPUT_DIR", r"C:\raw", ""])
            defaults_ws.append(["OUT_DIR", r"C:\out", ""])

            instruments_ws = workbook.create_sheet("Instruments")
            instruments_ws.append(["key", "component", "source"])
            instruments_ws.append(["power_kw", "dyno", "bench"])

            reporting_ws = workbook.create_sheet("Reporting_Rounding")
            reporting_ws.append(["key", "report_resolution", "report_digits", "rule", "notes"])
            reporting_ws.append(["power_kw", "0.1", "1", "round", ""])

            plots_ws = workbook.create_sheet("Plots")
            plots_ws.append(["enabled", "plot_type", "filename", "x_col", "y_col", "yerr_col"])
            plots_ws.append(["1", "yx", "power_vs_load.png", "Load_kW", "Power_kW", "off"])

            quality_ws = workbook.create_sheet("data quality assessment")
            quality_ws.append(["param", "value"])
            quality_ws.append(["MAX_DELTA_BETWEEN_SAMPLES_ms", 500.0])

            workbook.save(excel_path)

            bundle = bootstrap_text_config_from_excel(excel_path, text_dir)

            self.assertTrue(text_config_exists(text_dir))
            self.assertEqual(bundle.source_kind, "text")
            self.assertEqual(bundle.defaults["RAW_INPUT_DIR"], r"C:\raw")


if __name__ == "__main__":
    unittest.main()
