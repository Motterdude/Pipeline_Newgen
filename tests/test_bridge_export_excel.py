from __future__ import annotations

import importlib.util
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from _path import ROOT  # noqa: F401
from pipeline_newgen_rev1.bridges.legacy_runtime import ExportExcelBridgeStage
from pipeline_newgen_rev1.runtime.context import RuntimeContext


_MATPLOTLIB_AVAILABLE = importlib.util.find_spec("matplotlib") is not None
_LEGACY_REASON = "legacy monolith requires matplotlib (install with `pip install .[legacy]`)"


class ExportExcelBridgeTests(unittest.TestCase):
    def _ctx(self, *, output_dir: Path, final_table=None) -> RuntimeContext:
        ctx = RuntimeContext(project_root=output_dir)
        ctx.output_dir = output_dir
        ctx.final_table = final_table
        return ctx

    def test_skips_when_final_table_is_none(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            ctx = self._ctx(output_dir=out, final_table=None)
            ExportExcelBridgeStage().run(ctx)
            self.assertIsNone(ctx.lv_kpis_path)
            self.assertFalse((out / "lv_kpis_clean.xlsx").exists())

    @unittest.skipUnless(_MATPLOTLIB_AVAILABLE, _LEGACY_REASON)
    def test_writes_lv_kpis_when_frame_present(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            out = Path(tmp)
            frame = pd.DataFrame(
                {"Iteracao": [1, 2], "Sentido_Carga": ["sub", "des"], "Power_kW": [50.0, 75.0]}
            )
            ctx = self._ctx(output_dir=out, final_table=frame)
            ExportExcelBridgeStage().run(ctx)
            self.assertIsNotNone(ctx.lv_kpis_path)
            assert ctx.lv_kpis_path is not None  # narrow for mypy-style
            self.assertTrue(ctx.lv_kpis_path.exists())
            self.assertEqual(ctx.lv_kpis_path.name, "lv_kpis_clean.xlsx")

            wb = load_workbook(ctx.lv_kpis_path)
            sheet = wb[wb.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))
            self.assertEqual(rows[0], ("Iteracao", "Sentido_Carga", "Power_kW"))
            self.assertEqual(rows[1], (1, "sub", 50.0))
            self.assertEqual(rows[2], (2, "des", 75.0))
            wb.close()

    def test_raises_when_output_dir_missing(self) -> None:
        frame = pd.DataFrame({"x": [1]})
        ctx = RuntimeContext(project_root=Path("."))
        ctx.final_table = frame
        # output_dir remains None
        with self.assertRaises(RuntimeError):
            ExportExcelBridgeStage().run(ctx)


if __name__ == "__main__":
    unittest.main()
